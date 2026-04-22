from __future__ import annotations

from openpyxl import load_workbook

from data_engine.devtools.smoke_data import build_smoke_environment, build_temp_smoke_environment


def test_build_smoke_environment_creates_repo_style_data_and_workspaces(tmp_path):
    build_smoke_environment(root=tmp_path, workspace_ids=["example_workspace", "docs2"])

    assert (tmp_path / "data" / "Input" / "docs_flat" / "docs_flat_1.xlsx").exists() is True
    assert (tmp_path / "data2" / "Settings" / "single_watch.xlsx").exists() is True
    assert (tmp_path / "data2" / "Input" / "docs_parallel_12" / "docs_parallel_01.xlsx").exists() is True
    assert (tmp_path / "data2" / "Input" / "docs_parallel_12" / "docs_parallel_12.xlsx").exists() is True
    assert (tmp_path / "workspaces" / "example_workspace" / "flow_modules" / "example_manual.py").exists() is True
    starter_database_flow = tmp_path / "workspaces" / "example_workspace" / "flow_modules" / "example_database_dimensions.py"
    assert starter_database_flow.exists() is True
    starter_database_text = starter_database_flow.read_text(encoding="utf-8")
    assert 'context.database("docs/warehouse.duckdb")' in starter_database_text
    assert "build_dimension(" in starter_database_text
    assert (tmp_path / "workspaces" / "docs2" / "flow_modules" / "docs2_nb_manual.ipynb").exists() is True
    parallel_poll_flow = tmp_path / "workspaces" / "docs2" / "flow_modules" / "docs2_parallel_poll.py"
    parallel_schedule_flow = tmp_path / "workspaces" / "docs2" / "flow_modules" / "docs2_parallel_schedule.py"
    parallel_manual_flow = tmp_path / "workspaces" / "docs2" / "flow_modules" / "docs2_parallel_manual.py"
    assert parallel_poll_flow.exists() is True
    assert parallel_schedule_flow.exists() is True
    assert parallel_manual_flow.exists() is True
    assert "max_parallel=4" in parallel_poll_flow.read_text(encoding="utf-8")
    assert 'run_as="individual"' in parallel_schedule_flow.read_text(encoding="utf-8")


def test_build_temp_smoke_environment_preserves_live_suite_layout(tmp_path):
    build_temp_smoke_environment(temp_root=tmp_path, workspace_ids=["example_workspace"])

    assert (tmp_path / "app_root" / "config").is_dir() is True
    assert (tmp_path / "example_data" / "Input" / "docs_flat" / "docs_flat_1.xlsx").exists() is True
    assert (tmp_path / "data2" / "Settings" / "single_watch.xlsx").exists() is True
    assert (tmp_path / "workspaces" / "example_workspace" / "flow_modules" / "example_schedule.py").exists() is True


def test_build_smoke_environment_can_generate_wider_taller_claim_workbooks(tmp_path):
    build_smoke_environment(
        root=tmp_path,
        workspace_ids=["example_workspace"],
        rows_per_workbook=4,
        column_count=8,
    )

    workbook = load_workbook(
        tmp_path / "data" / "Input" / "docs_flat" / "docs_flat_1.xlsx",
        read_only=True,
        data_only=True,
    )
    docs_sheet = workbook["Docs"]
    summary_sheet = workbook["Summary"]
    try:
        header = next(docs_sheet.iter_rows(min_row=1, max_row=1, values_only=True))
        first_data_row = next(docs_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
        summary_row = next(summary_sheet.iter_rows(min_row=2, max_row=2, values_only=True))
    finally:
        workbook.close()

    assert len(header) == 8
    assert header[:5] == ("DCN", "Workflow", "Step FROM", "Step TO", "Employee ID")
    assert len(first_data_row) == 8
    assert summary_row == ("Docs", 4, 8)
