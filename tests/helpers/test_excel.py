from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.worksheet.table import Table
from openpyxl.worksheet.table import TableStyleInfo
import polars as pl
from polars.testing import assert_frame_equal
import pytest

from data_engine.helpers import ExcelSheet
from data_engine.helpers import compose_excel


def test_compose_excel_writes_multiple_named_sheets_and_tables(tmp_path: Path):
    target = tmp_path / "nested" / "report.xlsx"
    claims = pl.DataFrame({"claim_id": [1, 2], "workflow": ["Appeals", "Enrollment"]})
    summary = pl.DataFrame({"workflow": ["Appeals", "Enrollment"], "count": [1, 1]})

    returned_path = compose_excel(
        target,
        sheets=[
            ExcelSheet(name="Claims", df=claims, table_name="claims", freeze_panes="A2"),
            ExcelSheet(name="Summary", df=summary, table_name="workflow_summary"),
        ],
    )

    assert returned_path == target.resolve()
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), claims)
    assert_frame_equal(pl.read_excel(target, sheet_name="Summary"), summary)

    workbook = load_workbook(target)
    assert workbook.sheetnames == ["Claims", "Summary"]
    assert set(workbook["Claims"].tables) == {"claims"}
    assert set(workbook["Summary"].tables) == {"workflow_summary"}
    assert workbook["Claims"].freeze_panes == "A2"
    assert list(target.parent.glob(f".{target.name}.*.tmp.xlsx")) == []


def test_compose_excel_collects_lazy_frames(tmp_path: Path):
    target = tmp_path / "lazy.xlsx"
    frame = pl.DataFrame({"claim_id": [2, 1], "amount": [10, 20]})

    compose_excel(
        target,
        sheets=[
            ExcelSheet(
                name="Claims",
                df=frame.lazy().sort("claim_id"),
                table_name="claims",
            )
        ],
    )

    expected = pl.DataFrame({"claim_id": [1, 2], "amount": [20, 10]})
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), expected)


def test_compose_excel_replaces_existing_workbook_atomically(tmp_path: Path):
    target = tmp_path / "report.xlsx"
    old_frame = pl.DataFrame({"claim_id": [0]})
    new_frame = pl.DataFrame({"claim_id": [1, 2]})

    compose_excel(target, sheets=[ExcelSheet(name="Claims", df=old_frame, table_name="old_claims")])
    returned_path = compose_excel(target, sheets=[ExcelSheet(name="Claims", df=new_frame, table_name="new_claims")])

    assert returned_path == target.resolve()
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), new_frame)
    workbook = load_workbook(target)
    assert set(workbook["Claims"].tables) == {"new_claims"}
    assert list(target.parent.glob(f".{target.name}.*.tmp.xlsx")) == []


def test_compose_excel_applies_named_table_style_in_fresh_workbook(tmp_path: Path):
    target = tmp_path / "styled.xlsx"

    compose_excel(
        target,
        sheets=[
            ExcelSheet(
                name="Claims",
                df=pl.DataFrame({"claim_id": [1], "amount": [10.5]}),
                table_name="claims",
                table_style="TableStyleMedium9",
            )
        ],
    )

    table = load_workbook(target)["Claims"].tables["claims"]
    assert table.tableStyleInfo is not None
    assert table.tableStyleInfo.name == "TableStyleMedium9"


def test_compose_excel_forwards_fresh_workbook_write_options(tmp_path: Path):
    target = tmp_path / "formats.xlsx"

    compose_excel(
        target,
        sheets=[
            ExcelSheet(
                name="Claims",
                df=pl.DataFrame({"claim_id": [1], "amount": [10.5]}),
                table_name="claims",
                write_options={"hide_gridlines": True, "sheet_zoom": 125},
            )
        ],
    )

    worksheet = load_workbook(target)["Claims"]
    assert worksheet.sheet_view.showGridLines is False
    assert worksheet.sheet_view.zoomScale == 125


def test_compose_excel_template_path_can_update_same_target_and_preserve_other_sheets(tmp_path: Path):
    target = tmp_path / "template_report.xlsx"
    workbook = Workbook()
    data = workbook.active
    data.title = "Claims"
    data["A1"] = "old"
    pivot = workbook.create_sheet("Pivot")
    pivot["A1"] = "Keep this pivot-like sheet"
    workbook.save(target)

    frame = pl.DataFrame({"claim_id": [1, 2], "workflow": ["Appeals", "Enrollment"]})

    returned_path = compose_excel(
        target,
        sheets=[ExcelSheet(name="Claims", df=frame, table_name="claims", freeze_panes="A2")],
        template=target,
    )

    assert returned_path == target.resolve()
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), frame)
    updated = load_workbook(target)
    assert updated.sheetnames == ["Claims", "Pivot"]
    assert updated["Pivot"]["A1"].value == "Keep this pivot-like sheet"
    assert set(updated["Claims"].tables) == {"claims"}
    assert updated["Claims"].freeze_panes == "A2"
    assert list(target.parent.glob(f".{target.name}.*.tmp.xlsx")) == []


def test_compose_excel_template_applies_named_table_style(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "Claims"
    workbook.save(template)

    compose_excel(
        template,
        sheets=[
            ExcelSheet(
                name="Claims",
                df=pl.DataFrame({"claim_id": [1]}),
                table_name="claims",
                table_style="TableStyleMedium9",
            )
        ],
        template=template,
    )

    table = load_workbook(template)["Claims"].tables["claims"]
    assert table.tableStyleInfo is not None
    assert table.tableStyleInfo.name == "TableStyleMedium9"


def test_compose_excel_template_resizes_existing_table_to_new_frame_shape(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Claims"
    worksheet.append(["claim_id", "workflow"])
    worksheet.append([1, "Old"])
    table = Table(displayName="claims", ref="A1:B2")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True)
    worksheet.add_table(table)
    workbook.save(template)

    frame = pl.DataFrame(
        {
            "claim_id": [1, 2, 3, 4, 5, 6],
            "workflow": ["Appeals", "Enrollment", "Intake", "Review", "Archive", "Done"],
        }
    )

    compose_excel(
        template,
        sheets=[ExcelSheet(name="Claims", df=frame, table_name="claims")],
        template=template,
    )

    updated = load_workbook(template)
    assert updated["Claims"].tables["claims"].ref == "A1:B7"
    assert_frame_equal(pl.read_excel(template, sheet_name="Claims"), frame)


def test_compose_excel_template_updates_multiple_sheets_in_one_call(tmp_path: Path):
    template = tmp_path / "multi_template.xlsx"
    workbook = Workbook()
    workbook.active.title = "Claims"
    workbook.create_sheet("Summary")
    workbook.create_sheet("Pivot")
    workbook["Pivot"]["A1"] = "preserved"
    workbook.save(template)
    claims = pl.DataFrame({"claim_id": [1, 2], "workflow": ["Appeals", "Review"]})
    summary = pl.DataFrame({"workflow": ["Appeals", "Review"], "count": [1, 1]})

    compose_excel(
        template,
        sheets=[
            ExcelSheet(name="Claims", df=claims, table_name="claims"),
            ExcelSheet(name="Summary", df=summary, table_name="workflow_summary"),
        ],
        template=template,
    )

    updated = load_workbook(template)
    assert updated.sheetnames == ["Claims", "Summary", "Pivot"]
    assert updated["Pivot"]["A1"].value == "preserved"
    assert set(updated["Claims"].tables) == {"claims"}
    assert set(updated["Summary"].tables) == {"workflow_summary"}
    assert_frame_equal(pl.read_excel(template, sheet_name="Claims"), claims)
    assert_frame_equal(pl.read_excel(template, sheet_name="Summary"), summary)


def test_compose_excel_template_supports_cell_and_tuple_positions(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "Positioned"
    workbook.create_sheet("TuplePosition")
    workbook.save(template)

    compose_excel(
        template,
        sheets=[
            ExcelSheet(
                name="Positioned",
                df=pl.DataFrame({"claim_id": [1], "workflow": ["Appeals"]}),
                table_name="positioned",
                position="C3",
            ),
            ExcelSheet(
                name="TuplePosition",
                df=pl.DataFrame({"claim_id": [2], "workflow": ["Review"]}),
                table_name="tuple_positioned",
                position=(1, 2),
            ),
        ],
        template=template,
    )

    workbook = load_workbook(template)
    positioned = workbook["Positioned"]
    assert positioned["C3"].value == "claim_id"
    assert positioned["D3"].value == "workflow"
    assert positioned["C4"].value == 1
    assert positioned.tables["positioned"].ref == "C3:D4"
    tuple_positioned = workbook["TuplePosition"]
    assert tuple_positioned["C2"].value == "claim_id"
    assert tuple_positioned["D2"].value == "workflow"
    assert tuple_positioned["C3"].value == 2
    assert tuple_positioned.tables["tuple_positioned"].ref == "C2:D3"


def test_compose_excel_template_path_writes_output_without_changing_template(tmp_path: Path):
    template = tmp_path / "source_template.xlsx"
    target = tmp_path / "output" / "report.xlsx"
    workbook = Workbook()
    workbook.active.title = "Claims"
    workbook.active["A1"] = "old"
    notes = workbook.create_sheet("Notes")
    notes["A1"] = "preserved"
    workbook.save(template)

    frame = pl.DataFrame({"claim_id": [3]})

    compose_excel(
        target,
        sheets=[ExcelSheet(name="Claims", df=frame, table_name="claims")],
        template=template,
    )

    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), frame)
    output = load_workbook(target)
    assert output["Notes"]["A1"].value == "preserved"
    assert set(output["Claims"].tables) == {"claims"}
    unchanged_template = load_workbook(template)
    assert unchanged_template["Claims"]["A1"].value == "old"


def test_compose_excel_template_requires_existing_workbook(tmp_path: Path):
    with pytest.raises(ValueError, match="template workbook does not exist"):
        compose_excel(
            tmp_path / "report.xlsx",
            sheets=[ExcelSheet(name="Claims", df=pl.DataFrame({"claim_id": [1]}))],
            template=tmp_path / "missing_template.xlsx",
        )


def test_dataframe_namespace_composes_single_sheet_workbook(tmp_path: Path):
    target = tmp_path / "namespace.xlsx"
    frame = pl.DataFrame({"claim_id": [1], "workflow": ["Appeals"]})

    returned_path = frame.de.compose_excel(
        target,
        sheet_name="Claims",
        table_name="claims",
        table_style="TableStyleMedium9",
        freeze_panes="A2",
    )

    assert returned_path == target.resolve()
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), frame)
    worksheet = load_workbook(target)["Claims"]
    assert worksheet.tables["claims"].tableStyleInfo.name == "TableStyleMedium9"
    assert worksheet.freeze_panes == "A2"


def test_lazyframe_namespace_composes_single_sheet_workbook(tmp_path: Path):
    target = tmp_path / "lazy_namespace.xlsx"
    frame = pl.DataFrame({"claim_id": [2, 1], "workflow": ["Review", "Appeals"]})

    returned_path = frame.lazy().sort("claim_id").de.compose_excel(
        target,
        sheet_name="Claims",
        table_name="claims",
    )

    expected = pl.DataFrame({"claim_id": [1, 2], "workflow": ["Appeals", "Review"]})
    assert returned_path == target.resolve()
    assert_frame_equal(pl.read_excel(target, sheet_name="Claims"), expected)
    assert set(load_workbook(target)["Claims"].tables) == {"claims"}


def test_dataframe_namespace_compose_excel_supports_template_mode(tmp_path: Path):
    template = tmp_path / "template.xlsx"
    workbook = Workbook()
    workbook.active.title = "Claims"
    workbook.create_sheet("Pivot")
    workbook["Pivot"]["A1"] = "preserved"
    workbook.save(template)
    frame = pl.DataFrame({"claim_id": [1], "workflow": ["Appeals"]})

    frame.de.compose_excel(
        template,
        sheet_name="Claims",
        table_name="claims",
        template=template,
    )

    workbook = load_workbook(template)
    assert workbook["Pivot"]["A1"].value == "preserved"
    assert_frame_equal(pl.read_excel(template, sheet_name="Claims"), frame)


@pytest.mark.parametrize(
    ("sheets", "message"),
    [
        ([], "at least one"),
        ([ExcelSheet(name="", df=pl.DataFrame({"a": [1]}))], "sheet names must not be blank"),
        ([ExcelSheet(name="Bad/Name", df=pl.DataFrame({"a": [1]}))], "invalid Excel sheet-name"),
        (
            [
                ExcelSheet(name="Claims", df=pl.DataFrame({"a": [1]})),
                ExcelSheet(name="claims", df=pl.DataFrame({"a": [2]})),
            ],
            "sheet names must be unique",
        ),
        ([ExcelSheet(name="Claims", df=pl.DataFrame({"a": [1]}), table_name="bad table")], "table name"),
        ([ExcelSheet(name="Claims", df=pl.DataFrame({"a": [1]}), table_name="A1")], "cell reference"),
        (
            [
                ExcelSheet(name="Claims", df=pl.DataFrame({"a": [1]}), table_name="claims"),
                ExcelSheet(name="Summary", df=pl.DataFrame({"a": [2]}), table_name="CLAIMS"),
            ],
            "table names must be unique",
        ),
    ],
)
def test_compose_excel_validates_workbook_specs(tmp_path: Path, sheets, message: str):
    with pytest.raises(ValueError, match=message):
        compose_excel(tmp_path / "report.xlsx", sheets=sheets)


def test_compose_excel_rejects_non_polars_frames(tmp_path: Path):
    with pytest.raises(ValueError, match="DataFrame or LazyFrame"):
        compose_excel(
            tmp_path / "report.xlsx",
            sheets=[ExcelSheet(name="Claims", df={"claim_id": [1]})],  # type: ignore[arg-type]
        )

    assert list(tmp_path.glob(".report.xlsx.*.tmp.xlsx")) == []
