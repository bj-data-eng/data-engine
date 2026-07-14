"""Excel workbook composition helpers for Data Engine flow authoring."""

from __future__ import annotations

from collections.abc import Sequence
from dataclasses import dataclass, field
import os
from pathlib import Path
import re
import shutil
import time
from typing import Any
from uuid import uuid4

import polars as pl
import xlsxwriter

PathLike = str | os.PathLike[str]
ExcelFrame = pl.DataFrame | pl.LazyFrame

_INVALID_SHEET_CHARS = set("[]:*?/\\")
_CELL_REFERENCE_RE = re.compile(r"^[A-Za-z]{1,3}[1-9][0-9]*$")
_TABLE_NAME_RE = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")


@dataclass(frozen=True)
class ExcelSheet:
    """Specification for one dataframe-backed worksheet in an Excel workbook.

    Attributes
    ----------
    name : str
        Worksheet name. Names must follow Excel's normal sheet-name limits.
    df : ExcelFrame
        Dataframe to write. Lazy frames are collected when the workbook is
        composed.
    table_name : str | None
        Optional Excel table name for the written dataframe.
    position : str | tuple[int, int]
        Top-left cell where the dataframe table should be written.
    table_style : str | dict[str, Any] | None
        Optional table style forwarded to ``pl.DataFrame.write_excel``.
    autofit : bool
        Whether Polars should autofit columns after writing.
    autofilter : bool
        Whether the generated table should include filter controls.
    freeze_panes : object | None
        Optional freeze-pane setting forwarded to ``pl.DataFrame.write_excel``.
    write_options : dict[str, Any]
        Additional keyword options forwarded to ``pl.DataFrame.write_excel``.
    """

    name: str
    df: ExcelFrame
    table_name: str | None = None
    position: str | tuple[int, int] = "A1"
    table_style: str | dict[str, Any] | None = None
    autofit: bool = True
    autofilter: bool = True
    freeze_panes: object | None = None
    write_options: dict[str, Any] = field(default_factory=dict)


def compose_excel(path: PathLike, sheets: Sequence[ExcelSheet], *, template: PathLike | None = None) -> Path:
    """Compose one Excel workbook from dataframe-backed sheet specifications.

    The workbook is written to a same-directory temporary file and then moved
    into place with ``os.replace``. Each sheet is written through Polars'
    Excel writer, so sheet-level options map directly to
    ``pl.DataFrame.write_excel`` behavior.

    Parameters
    ----------
    path : PathLike
        Target Excel workbook path.
    sheets : Sequence[ExcelSheet]
        One or more worksheet specifications to write.
    template : PathLike | None
        Optional template workbook path. When provided, the template workbook is
        copied to a temporary file, the requested sheets are updated, and the
        composed copy atomically replaces ``path``. Existing values and tables
        are replaced on each updated worksheet. Merged ranges that intersect the
        dataframe or table output are unmerged; unrelated merges and cell
        formatting are preserved.

    Returns
    -------
    Path
        Absolute target path that was replaced.

    Raises
    ------
    Exception
        If workbook validation, dataframe collection, Excel writing, or atomic
        replacement fails.

    Examples
    --------
    .. code-block:: python

        import polars as pl

        from data_engine.helpers import ExcelSheet, compose_excel

        compose_excel(
            "workspaces/example/output/report.xlsx",
            sheets=[
                ExcelSheet(
                    name="Claims",
                    df=pl.DataFrame({"claim_id": [1, 2]}),
                    table_name="claims",
                    freeze_panes="A2",
                ),
            ],
        )
    """
    normalized_sheets = tuple(sheets)
    _validate_sheets(normalized_sheets)
    target_path = Path(path).expanduser().resolve()
    target_path.parent.mkdir(parents=True, exist_ok=True)
    temporary_path = target_path.with_name(f".{target_path.name}.{uuid4().hex}.tmp.xlsx")
    if template is not None:
        return _compose_template_excel(
            target_path,
            temporary_path,
            sheets=normalized_sheets,
            template=template,
        )
    workbook = xlsxwriter.Workbook(temporary_path)
    try:
        for sheet in normalized_sheets:
            frame = _collect_frame(sheet.df)
            options = dict(sheet.write_options)
            frame.write_excel(
                workbook,
                worksheet=sheet.name,
                position=sheet.position,
                table_name=sheet.table_name,
                table_style=sheet.table_style,
                autofit=sheet.autofit,
                autofilter=sheet.autofilter,
                freeze_panes=sheet.freeze_panes,
                **options,
            )
        workbook.close()
        _replace_atomic(temporary_path, target_path)
    except Exception:
        try:
            workbook.close()
        except Exception:
            pass
        _remove_temporary_file(temporary_path)
        raise
    return target_path


def _compose_template_excel(
    target_path: Path,
    temporary_path: Path,
    *,
    sheets: tuple[ExcelSheet, ...],
    template: PathLike,
) -> Path:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ImportError("compose_excel(template=...) requires openpyxl.") from exc

    template_path = Path(template).expanduser().resolve()
    if not template_path.is_file():
        raise ValueError(f"template workbook does not exist: {template_path}")
    try:
        shutil.copy2(template_path, temporary_path)
        workbook = load_workbook(temporary_path)
        for sheet in sheets:
            frame = _collect_frame(sheet.df)
            _replace_template_sheet(workbook, sheet, frame)
        workbook.save(temporary_path)
        _replace_atomic(temporary_path, target_path)
    except Exception:
        _remove_temporary_file(temporary_path)
        raise
    return target_path


def _replace_template_sheet(workbook: Any, sheet: ExcelSheet, frame: pl.DataFrame) -> None:
    if sheet.name in workbook.sheetnames:
        worksheet = workbook[sheet.name]
    else:
        worksheet = workbook.create_sheet(sheet.name)
    start_row, start_column = _position_to_one_based(sheet.position)
    replacement_bounds: tuple[int, int, int, int] | None = None
    if frame.columns:
        data_row_count = max(frame.height, 1) if sheet.table_name is not None else frame.height
        replacement_bounds = (
            start_row,
            start_column,
            start_row + data_row_count,
            start_column + len(frame.columns) - 1,
        )
    _clear_worksheet_data(worksheet, replacement_bounds=replacement_bounds)
    for column_offset, column_name in enumerate(frame.columns):
        worksheet.cell(row=start_row, column=start_column + column_offset, value=column_name)
    for row_offset, values in enumerate(frame.iter_rows(), start=1):
        for column_offset, value in enumerate(values):
            worksheet.cell(row=start_row + row_offset, column=start_column + column_offset, value=value)
    if sheet.table_name is not None and frame.columns:
        _add_openpyxl_table(
            worksheet,
            table_name=sheet.table_name,
            start_row=start_row,
            start_column=start_column,
            column_count=len(frame.columns),
            row_count=frame.height,
            table_style=sheet.table_style,
        )
    if sheet.freeze_panes is not None:
        worksheet.freeze_panes = sheet.freeze_panes


def _clear_worksheet_data(
    worksheet: Any,
    *,
    replacement_bounds: tuple[int, int, int, int] | None,
) -> None:
    from openpyxl.cell.cell import MergedCell

    for table_name in list(worksheet.tables):
        del worksheet.tables[table_name]
    if replacement_bounds is not None:
        start_row, start_column, end_row, end_column = replacement_bounds
        overlapping_merged_ranges = [
            merged_range
            for merged_range in worksheet.merged_cells.ranges
            if (
                merged_range.min_row <= end_row
                and merged_range.max_row >= start_row
                and merged_range.min_col <= end_column
                and merged_range.max_col >= start_column
            )
        ]
        for merged_range in overlapping_merged_ranges:
            worksheet.unmerge_cells(merged_range.coord)

    # ``iter_rows`` materializes every coordinate in a sparse sheet's bounding
    # box. Walking openpyxl's existing cell store keeps clearing linear in the
    # cells the template already contains.
    for cell in worksheet._cells.values():
        if not isinstance(cell, MergedCell):
            cell.value = None


def _add_openpyxl_table(
    worksheet: Any,
    *,
    table_name: str,
    start_row: int,
    start_column: int,
    column_count: int,
    row_count: int,
    table_style: str | dict[str, Any] | None,
) -> None:
    from openpyxl.utils import get_column_letter
    from openpyxl.worksheet.table import Table, TableStyleInfo

    end_row = start_row + max(row_count, 1)
    end_column = start_column + column_count - 1
    ref = (
        f"{get_column_letter(start_column)}{start_row}:"
        f"{get_column_letter(end_column)}{end_row}"
    )
    table = Table(displayName=table_name, ref=ref)
    style_name = table_style if isinstance(table_style, str) else "TableStyleMedium2"
    table.tableStyleInfo = TableStyleInfo(
        name=style_name,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=True,
        showColumnStripes=False,
    )
    worksheet.add_table(table)


def _position_to_one_based(position: str | tuple[int, int]) -> tuple[int, int]:
    if isinstance(position, str):
        from openpyxl.utils.cell import coordinate_to_tuple

        try:
            return coordinate_to_tuple(position)
        except ValueError as exc:
            raise ValueError(f"position must be an Excel cell reference or a row/column tuple: {position!r}") from exc
    if len(position) != 2:
        raise ValueError("position tuple must contain row and column.")
    row, column = position
    if row < 0 or column < 0:
        raise ValueError("position tuple values must be zero or greater.")
    return row + 1, column + 1


def _collect_frame(df: ExcelFrame) -> pl.DataFrame:
    if isinstance(df, pl.DataFrame):
        return df
    if isinstance(df, pl.LazyFrame):
        return df.collect()
    raise ValueError("ExcelSheet.df must be a Polars DataFrame or LazyFrame.")


def _validate_sheets(sheets: tuple[ExcelSheet, ...]) -> None:
    if not sheets:
        raise ValueError("sheets must contain at least one ExcelSheet.")
    seen_sheet_names: set[str] = set()
    seen_table_names: set[str] = set()
    for sheet in sheets:
        if not isinstance(sheet, ExcelSheet):
            raise ValueError("sheets must contain only ExcelSheet values.")
        normalized_sheet_name = _validate_sheet_name(sheet.name)
        if normalized_sheet_name in seen_sheet_names:
            raise ValueError(f"sheet names must be unique: {sheet.name!r}")
        seen_sheet_names.add(normalized_sheet_name)
        if sheet.table_name is not None:
            normalized_table_name = _validate_table_name(sheet.table_name)
            if normalized_table_name in seen_table_names:
                raise ValueError(f"table names must be unique: {sheet.table_name!r}")
            seen_table_names.add(normalized_table_name)


def _validate_sheet_name(name: str) -> str:
    normalized = name.strip()
    if not normalized:
        raise ValueError("sheet names must not be blank.")
    if len(normalized) > 31:
        raise ValueError(f"sheet name {name!r} must be 31 characters or fewer.")
    if any(char in _INVALID_SHEET_CHARS for char in normalized):
        raise ValueError(f"sheet name {name!r} contains an invalid Excel sheet-name character.")
    return normalized.casefold()


def _validate_table_name(name: str) -> str:
    normalized = name.strip()
    if not normalized:
        raise ValueError("table names must not be blank.")
    if len(normalized) > 255:
        raise ValueError(f"table name {name!r} must be 255 characters or fewer.")
    if not _TABLE_NAME_RE.fullmatch(normalized):
        raise ValueError(
            f"table name {name!r} must start with a letter or underscore and contain only letters, numbers, and underscores."
        )
    if _CELL_REFERENCE_RE.fullmatch(normalized):
        raise ValueError(f"table name {name!r} must not look like an Excel cell reference.")
    return normalized.casefold()


def _replace_atomic(source_path: Path, target_path: Path) -> None:
    backoff_seconds = (0.0, 0.02, 0.05, 0.1, 0.2)
    last_error: BaseException | None = None
    for delay_seconds in backoff_seconds:
        if delay_seconds > 0.0:
            time.sleep(delay_seconds)
        try:
            os.replace(source_path, target_path)
            return
        except PermissionError as exc:
            if getattr(exc, "winerror", None) != 5:
                raise
            last_error = exc
            continue
    if last_error is not None:
        raise last_error
    os.replace(source_path, target_path)


def _remove_temporary_file(path: Path) -> None:
    try:
        path.unlink()
    except FileNotFoundError:
        pass
    except OSError:
        pass


__all__ = ["ExcelSheet", "compose_excel"]
