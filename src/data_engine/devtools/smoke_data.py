"""Reusable smoke data and starter-workspace generators."""

from __future__ import annotations

import argparse
import json
from pathlib import Path

from openpyxl import Workbook


DEFAULT_WORKSPACE_IDS = ("example_workspace", "claims2")
_BASE_CLAIMS_COLUMNS = (
    "DCN",
    "Workflow",
    "Step FROM",
    "Step TO",
    "Employee ID",
)


def build_temp_smoke_environment(*, temp_root: Path, workspace_ids: list[str]) -> None:
    """Build the temporary smoke-test environment used by the live runtime suite."""
    build_smoke_environment(
        root=temp_root,
        workspace_ids=workspace_ids,
        primary_data_dir_name="example_data",
        secondary_data_dir_name="data2",
        create_app_root=True,
    )


def build_smoke_environment(
    *,
    root: Path,
    workspace_ids: list[str],
    primary_data_dir_name: str = "data",
    secondary_data_dir_name: str = "data2",
    create_app_root: bool = False,
    rows_per_workbook: int = 2,
    column_count: int = len(_BASE_CLAIMS_COLUMNS),
) -> None:
    """Generate starter data roots plus authored workspaces under one root."""
    root = Path(root).expanduser().resolve()
    workspace_collection_root = root / "workspaces"
    primary_data_root = root / primary_data_dir_name
    secondary_data_root = root / secondary_data_dir_name

    if create_app_root:
        (root / "app_root" / "config").mkdir(parents=True, exist_ok=True)
    workspace_collection_root.mkdir(parents=True, exist_ok=True)
    create_smoke_data_root(
        primary_data_root,
        rows_per_workbook=rows_per_workbook,
        column_count=column_count,
    )
    create_smoke_data_root(
        secondary_data_root,
        rows_per_workbook=rows_per_workbook,
        column_count=column_count,
    )

    for workspace_id in workspace_ids:
        target_workspace = workspace_collection_root / workspace_id
        (target_workspace / "flow_modules").mkdir(parents=True, exist_ok=True)
        data_folder_name = secondary_data_dir_name if workspace_id.endswith("2") else primary_data_dir_name
        create_python_flow_modules(target_workspace, workspace_id=workspace_id, data_folder_name=data_folder_name)
        create_notebook_flow_modules(target_workspace, workspace_id=workspace_id, data_folder_name=data_folder_name)


def create_python_flow_modules(target_workspace: Path, *, workspace_id: str, data_folder_name: str) -> None:
    """Write starter Python-authored flow modules into one workspace."""
    del workspace_id
    flow_dir = target_workspace / "flow_modules"
    write_text_file(flow_dir / "example_mirror.py", _python_poll_source(data_folder_name=data_folder_name))
    write_text_file(flow_dir / "example_schedule.py", _python_schedule_source(data_folder_name=data_folder_name))
    write_text_file(flow_dir / "example_manual.py", _python_manual_source(data_folder_name=data_folder_name))
    write_text_file(
        flow_dir / "example_database_dimensions.py",
        _python_database_dimensions_source(data_folder_name=data_folder_name),
    )


def create_notebook_flow_modules(target_workspace: Path, *, workspace_id: str, data_folder_name: str) -> None:
    """Write starter notebook-authored flow modules into one workspace."""
    flow_dir = target_workspace / "flow_modules"
    create_notebook_flow_module(
        flow_dir / f"{workspace_id}_nb_poll.ipynb",
        _poll_notebook_source(workspace_id=workspace_id, data_folder_name=data_folder_name),
    )
    create_notebook_flow_module(
        flow_dir / f"{workspace_id}_nb_schedule.ipynb",
        _schedule_notebook_source(workspace_id=workspace_id, data_folder_name=data_folder_name),
    )
    create_notebook_flow_module(
        flow_dir / f"{workspace_id}_nb_manual.ipynb",
        _manual_notebook_source(workspace_id=workspace_id, data_folder_name=data_folder_name),
    )


def create_smoke_data_root(
    data_root: Path,
    *,
    rows_per_workbook: int = 2,
    column_count: int = len(_BASE_CLAIMS_COLUMNS),
) -> None:
    """Generate one smoke-data root with starter Excel inputs and output folders."""
    claims_flat = data_root / "Input" / "claims_flat"
    settings_dir = data_root / "Settings"
    output_dir = data_root / "Output"
    database_dir = data_root / "databases"
    claims_flat.mkdir(parents=True, exist_ok=True)
    settings_dir.mkdir(parents=True, exist_ok=True)
    output_dir.mkdir(parents=True, exist_ok=True)
    database_dir.mkdir(parents=True, exist_ok=True)
    if rows_per_workbook <= 0:
        raise ValueError("rows_per_workbook must be positive.")
    if column_count < len(_BASE_CLAIMS_COLUMNS):
        raise ValueError(
            f"column_count must be at least {len(_BASE_CLAIMS_COLUMNS)} to preserve the starter claims schema."
        )

    for index in range(1, 4):
        workbook_path = claims_flat / f"claims_flat_{index}.xlsx"
        _write_claims_workbook(
            workbook_path,
            workbook_index=index,
            rows_per_workbook=rows_per_workbook,
            column_count=column_count,
        )

    settings_path = settings_dir / "single_watch.xlsx"
    _write_claims_workbook(
        settings_path,
        workbook_index=0,
        rows_per_workbook=rows_per_workbook,
        column_count=column_count,
    )


def _claims_headers(*, column_count: int) -> tuple[str, ...]:
    extra_count = column_count - len(_BASE_CLAIMS_COLUMNS)
    extra_columns = tuple(f"Attribute {index:02d}" for index in range(1, extra_count + 1))
    return (*_BASE_CLAIMS_COLUMNS, *extra_columns)


def _claims_row(*, workbook_index: int, row_number: int, column_count: int) -> tuple[object, ...]:
    workflow = ("Claims", "Appeals", "Enrollment", "Billing")[row_number % 4]
    step_from = ("Receive", "Review", "Triage", "Audit")[row_number % 4]
    step_to = ("Review", "Approve", "Finalize", "Queue")[row_number % 4]
    base_values: list[object] = [
        f"{workbook_index:02d}{row_number:08d}",
        workflow,
        step_from,
        step_to,
        f"E-{((row_number + workbook_index) % 9000) + 1:04d}",
    ]
    for extra_index in range(column_count - len(_BASE_CLAIMS_COLUMNS)):
        if extra_index % 5 == 0:
            value: object = f"segment-{workbook_index}-{(row_number + extra_index) % 37:02d}"
        elif extra_index % 5 == 1:
            value = (row_number * (extra_index + 3)) % 100000
        elif extra_index % 5 == 2:
            value = round(((row_number % 1000) / 1000) + workbook_index + (extra_index / 100), 4)
        elif extra_index % 5 == 3:
            value = "Y" if (row_number + extra_index + workbook_index) % 2 == 0 else "N"
        else:
            value = f"group-{((row_number // 1000) + workbook_index + extra_index) % 120:03d}"
        base_values.append(value)
    return tuple(base_values)


def _write_claims_workbook(
    path: Path,
    *,
    workbook_index: int,
    rows_per_workbook: int,
    column_count: int,
) -> None:
    workbook = Workbook(write_only=True)
    claims_sheet = workbook.create_sheet("Claims")
    claims_sheet.append(_claims_headers(column_count=column_count))
    for row_number in range(1, rows_per_workbook + 1):
        claims_sheet.append(_claims_row(workbook_index=workbook_index, row_number=row_number, column_count=column_count))

    summary_sheet = workbook.create_sheet("Summary")
    summary_sheet.append(("Sheet", "Rows", "Columns"))
    summary_sheet.append(("Claims", rows_per_workbook, column_count))
    workbook.save(path)


def create_notebook_flow_module(path: Path, source_text: str) -> None:
    """Write one single-cell notebook flow module."""
    notebook = {
        "cells": [
            {
                "cell_type": "code",
                "execution_count": None,
                "metadata": {},
                "outputs": [],
                "source": [line if line.endswith("\n") else f"{line}\n" for line in source_text.splitlines()],
            }
        ],
        "metadata": {
            "kernelspec": {"display_name": "Python 3", "language": "python", "name": "python3"},
            "language_info": {"name": "python"},
        },
        "nbformat": 4,
        "nbformat_minor": 5,
    }
    path.write_text(json.dumps(notebook, indent=2), encoding="utf-8")


def write_text_file(path: Path, text: str) -> None:
    """Write one text file with a guaranteed trailing newline."""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(text.rstrip() + "\n", encoding="utf-8")


def build_parser() -> argparse.ArgumentParser:
    """Build the CLI parser for generating local smoke data."""
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument(
        "--root",
        type=Path,
        default=Path.cwd(),
        help="Root directory that will receive workspaces and data folders.",
    )
    parser.add_argument(
        "--workspace-id",
        dest="workspace_ids",
        action="append",
        help="Workspace id to generate. Defaults to example_workspace and claims2.",
    )
    parser.add_argument(
        "--primary-data-dir",
        default="data",
        help="Directory name for the primary data root.",
    )
    parser.add_argument(
        "--secondary-data-dir",
        default="data2",
        help="Directory name for the secondary data root used by *2 workspaces.",
    )
    parser.add_argument(
        "--with-app-root",
        action="store_true",
        help="Also create app_root/config under the target root.",
    )
    parser.add_argument(
        "--rows-per-workbook",
        type=int,
        default=2,
        help="Number of claim rows to generate in each workbook.",
    )
    parser.add_argument(
        "--column-count",
        type=int,
        default=len(_BASE_CLAIMS_COLUMNS),
        help="Total number of columns to generate in each claims workbook.",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    """Generate starter smoke data and workspace folders."""
    args = build_parser().parse_args(argv)
    workspace_ids = args.workspace_ids or list(DEFAULT_WORKSPACE_IDS)
    build_smoke_environment(
        root=args.root,
        workspace_ids=workspace_ids,
        primary_data_dir_name=args.primary_data_dir,
        secondary_data_dir_name=args.secondary_data_dir,
        create_app_root=args.with_app_root,
        rows_per_workbook=args.rows_per_workbook,
        column_count=args.column_count,
    )
    print(f"Generated workspaces in {args.root.expanduser().resolve() / 'workspaces'}")
    print(f"Generated primary data in {args.root.expanduser().resolve() / args.primary_data_dir}")
    print(f"Generated secondary data in {args.root.expanduser().resolve() / args.secondary_data_dir}")
    return 0


def _python_poll_source(*, data_folder_name: str) -> str:
    return f"""from __future__ import annotations

import polars as pl

from data_engine import Flow

DESCRIPTION = "Reads workbook inputs and writes mirrored parquet outputs."


def read_claims(context):
    return pl.read_excel(context.source.path)


def write_target(context):
    output = context.mirror.with_suffix(".parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="example_mirror", group="Claims")
        .watch(
            mode="poll",
            source="../../../{data_folder_name}/Input/claims_flat",
            interval="5s",
            extensions=[".xlsx", ".xls", ".xlsm"],
            settle=1,
        )
        .mirror(root="../../../{data_folder_name}/Output/example_mirror")
        .step(read_claims, label="Read Excel")
        .step(write_target, label="Write Parquet")
    )
"""


def _python_schedule_source(*, data_folder_name: str) -> str:
    return f"""from __future__ import annotations

import polars as pl

from data_engine import Flow

DESCRIPTION = "Reads the settings workbook on a schedule and writes parquet output."


def read_settings(context):
    first_sheet = pl.read_excel(context.source.path, sheet_id=1)
    claims_sheet = pl.read_excel(context.source.path, sheet_name="Claims")
    return pl.concat((first_sheet, claims_sheet), how="vertical_relaxed")


def write_example_schedule(context):
    output = context.mirror.with_suffix(".parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="example_schedule", group="Settings")
        .watch(mode="schedule", run_as="batch", interval="15m", source="../../../{data_folder_name}/Settings/single_watch.xlsx")
        .mirror(root="../../../{data_folder_name}/Output/example_schedule")
        .step(read_settings, save_as="settings_df", label="Read Excel")
        .step(write_example_schedule, use="settings_df", label="Write Parquet")
    )
"""


def _python_manual_source(*, data_folder_name: str) -> str:
    return f"""import polars as pl

from data_engine import Flow

DESCRIPTION = "Manual starter flow that loads one workbook on demand and writes one parquet output."


def read_claims(file_ref):
    return pl.read_excel(file_ref.path)


def combine_claims(context):
    return pl.concat(context.current, how="vertical_relaxed")


def write_target(context):
    output = context.mirror.file("example_manual.parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="example_manual", group="Manual")
        .mirror(root="../../../{data_folder_name}/Output/example_manual")
        .collect([".xlsx"], root="../../../{data_folder_name}/Input/claims_flat", label="Collect Files")
        .map(read_claims, label="Read Excel")
        .step(combine_claims, label="Combine Claims")
        .step(write_target, label="Write Parquet")
    )
"""


def _python_database_dimensions_source(*, data_folder_name: str) -> str:
    return f"""import polars as pl

from data_engine import Flow
from data_engine.helpers.duckdb import attach_dimension
from data_engine.helpers.duckdb import build_dimension
from data_engine.helpers.duckdb import read_table
from data_engine.helpers.duckdb import replace_table

DESCRIPTION = "Manual starter flow that writes claims into a workspace-local DuckDB database and builds a surrogate-key dimension."


def read_claims(file_ref):
    return pl.read_excel(file_ref.path).rename(
        {{
            "DCN": "claim_id",
            "Workflow": "workflow",
            "Step FROM": "step_from",
            "Step TO": "step_to",
            "Employee ID": "employee_id",
        }}
    )


def combine_claims(context):
    frames = tuple(context.current)
    if not frames:
        return pl.DataFrame()
    return pl.concat(frames, how="vertical_relaxed").unique(maintain_order=True)


def persist_claim_warehouse(context):
    claims_df = context.current
    db_path = context.database("claims/warehouse.duckdb")

    replace_table(db_path, "stage.claims_flat", df=claims_df, return_df=False)
    build_dimension(
        db_path,
        "mart.dim_employee_workflow",
        df=claims_df.select(["employee_id", "workflow"]).unique(maintain_order=True),
        key_column="employee_workflow_key",
        return_df=False,
    )
    keyed_claims = attach_dimension(
        db_path,
        "mart.dim_employee_workflow",
        df=claims_df,
        on=["employee_id", "workflow"],
        key_column="employee_workflow_key",
    )
    replace_table(db_path, "mart.fact_claim", df=keyed_claims, return_df=False)
    return read_table(db_path, "mart.fact_claim")


def write_target(context):
    output = context.mirror.file("claims_dimension_snapshot.parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="example_database_dimensions", group="Warehouse")
        .mirror(root="../../../{data_folder_name}/Output/example_database_dimensions")
        .collect([".xlsx"], root="../../../{data_folder_name}/Input/claims_flat", label="Collect Files")
        .map(read_claims, label="Read Excel")
        .step(combine_claims, label="Combine Claims")
        .step(persist_claim_warehouse, label="Build DuckDB Dimension")
        .step(write_target, label="Write Snapshot")
    )
"""


def _poll_notebook_source(*, workspace_id: str, data_folder_name: str) -> str:
    return f"""from __future__ import annotations

import polars as pl

from data_engine import Flow

DESCRIPTION = "Notebook-authored poll flow for smoke testing."


def read_claims(context):
    return pl.read_excel(context.source.path)


def write_target(context):
    output = context.mirror.with_suffix(".parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="{workspace_id}_nb_poll", group="Notebook")
        .watch(
            mode="poll",
            source="../../../{data_folder_name}/Input/claims_flat",
            interval="5s",
            extensions=[".xlsx", ".xls", ".xlsm"],
            settle=1,
        )
        .mirror(root="../../../{data_folder_name}/Output/{workspace_id}_nb_poll")
        .step(read_claims, label="Read Excel")
        .step(write_target, label="Write Parquet")
    )
"""


def _schedule_notebook_source(*, workspace_id: str, data_folder_name: str) -> str:
    return f"""from __future__ import annotations

import polars as pl

from data_engine import Flow

DESCRIPTION = "Notebook-authored schedule flow for smoke testing."


def read_settings(context):
    first_sheet = pl.read_excel(context.source.path, sheet_id=1)
    claims_sheet = pl.read_excel(context.source.path, sheet_name="Claims")
    return pl.concat((first_sheet, claims_sheet), how="vertical_relaxed")


def write_example_schedule(context):
    output = context.mirror.with_suffix(".parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="{workspace_id}_nb_schedule", group="Notebook")
        .watch(mode="schedule", run_as="batch", interval="15m", source="../../../{data_folder_name}/Settings/single_watch.xlsx")
        .mirror(root="../../../{data_folder_name}/Output/{workspace_id}_nb_schedule")
        .step(read_settings, save_as="settings_df", label="Read Excel")
        .step(write_example_schedule, use="settings_df", label="Write Parquet")
    )
"""


def _manual_notebook_source(*, workspace_id: str, data_folder_name: str) -> str:
    return f"""import polars as pl

from data_engine import Flow

DESCRIPTION = "Notebook-authored manual flow for smoke testing."


def read_claims(file_ref):
    return pl.read_excel(file_ref.path)


def combine_claims(context):
    frames = tuple(context.current)
    if not frames:
        return pl.DataFrame()
    return pl.concat(frames, how="vertical_relaxed")


def write_target(context):
    output = context.mirror.file("manual_claims.parquet")
    context.current.write_parquet(output)
    return output


def build():
    return (
        Flow(name="{workspace_id}_nb_manual", group="Notebook")
        .mirror(root="../../../{data_folder_name}/Output/{workspace_id}_nb_manual")
        .collect([".xlsx"], root="../../../{data_folder_name}/Input/claims_flat", label="Collect Files")
        .map(read_claims, label="Read Excel")
        .step(combine_claims, label="Combine Claims")
        .step(write_target, label="Write Parquet")
    )
"""


__all__ = [
    "DEFAULT_WORKSPACE_IDS",
    "build_parser",
    "build_smoke_environment",
    "build_temp_smoke_environment",
    "create_notebook_flow_modules",
    "create_python_flow_modules",
    "create_smoke_data_root",
    "main",
]
